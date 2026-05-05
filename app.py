import os
import io
import re
from dotenv import load_dotenv
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, Room, Receptionist, Checklist
from rooms_data import get_all_rooms, sort_buildings
from sqlalchemy import func, cast, Date

load_dotenv()

SANTIAGO_TZ = ZoneInfo('America/Santiago')

def now_santiago():
    return datetime.now(SANTIAGO_TZ).replace(tzinfo=None)

app = Flask(__name__)

database_url = os.environ.get('DATABASE_URL', 'sqlite:///camp_checklist.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')

DELETE_PASSWORD = os.environ.get('DELETE_PASSWORD', 'admin2026')
MODULE_NAME_RE = re.compile(r'^[A-Z0-9\-]+$')

OKX_FIELDS = [
    'luz_central', 'sensor', 'cobertores', 'cambio_sabanas',
    'velador', 'almohada', 'extractor', 'estufa', 'basurero',
    'humidificador', 'cortina', 'blackout', 'aseo_general', 'closet'
]

OKX_LABELS = {
    'luz_central': 'Luz Central',
    'sensor': 'Sensor',
    'cobertores': 'Cobertores',
    'cambio_sabanas': 'Cambio Sabanas',
    'velador': 'Velador',
    'almohada': 'Almohada',
    'extractor': 'Extractor',
    'estufa': 'Estufa',
    'basurero': 'Basurero',
    'humidificador': 'Humidificador',
    'cortina': 'Cortina',
    'blackout': 'Blackout',
    'aseo_general': 'Aseo General',
    'closet': 'Closet',
}

db.init_app(app)


def seed_default_rooms_if_empty():
    if Room.query.first() is not None:
        return
    for code, building in get_all_rooms():
        db.session.add(Room(code=code, building=building))
    db.session.commit()


def normalize_module_name(name):
    return re.sub(r'\s+', '', (name or '').upper())


def get_available_modules():
    rows = db.session.query(
        Room.building,
        func.count(Room.id).label('room_count')
    ).group_by(Room.building).all()
    counts = {row.building: row.room_count for row in rows}
    ordered = sort_buildings(list(counts.keys()))
    return [{'building': b, 'room_count': counts[b]} for b in ordered]


def get_dashboard_range():
    selected_date = (request.args.get('selected_date') or '').strip()
    days = request.args.get('days', 7, type=int)
    if selected_date:
        try:
            start_date = datetime.strptime(selected_date, '%Y-%m-%d')
            end_date = start_date + timedelta(days=1)
            return {'start_date': start_date, 'end_date': end_date,
                    'selected_date': selected_date, 'days': None, 'label': selected_date}
        except ValueError:
            pass
    start_date = now_santiago() - timedelta(days=days)
    return {'start_date': start_date, 'end_date': None,
            'selected_date': '', 'days': days, 'label': f'Ultimos {days} dia(s)'}


with app.app_context():
    db.create_all()
    seed_default_rooms_if_empty()


# ── Pages ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    receptionists = Receptionist.query.filter_by(active=True).order_by(Receptionist.name).all()
    modules = get_available_modules()
    return render_template('index.html', receptionists=receptionists, modules=modules)


@app.route('/checklist/<building>')
def checklist_page(building):
    receptionist_id = request.args.get('receptionist_id', type=int)
    if not receptionist_id:
        return redirect(url_for('index'))
    receptionist = Receptionist.query.get_or_404(receptionist_id)
    rooms = Room.query.filter_by(building=building).order_by(Room.code).all()
    return render_template('checklist.html', rooms=rooms, building=building, receptionist=receptionist)


@app.route('/dashboard')
def dashboard_page():
    return render_template('dashboard.html')


@app.route('/receptionists')
def receptionists_page():
    receptionists = Receptionist.query.order_by(Receptionist.name).all()
    return render_template('receptionists.html', receptionists=receptionists)


@app.route('/history')
def history_page():
    return render_template('history.html')


# ── API: Modules ───────────────────────────────────────────────────────────

@app.route('/api/modules', methods=['GET'])
def api_get_modules():
    return jsonify(get_available_modules())


@app.route('/api/modules', methods=['POST'])
def api_create_module():
    data = request.get_json(silent=True) or {}
    building = normalize_module_name(data.get('building'))
    room_count = data.get('room_count', 0)
    try:
        room_count = int(room_count)
    except (TypeError, ValueError):
        room_count = 0
    if not building:
        return jsonify({'error': 'El nombre del modulo es requerido'}), 400
    if not MODULE_NAME_RE.fullmatch(building):
        return jsonify({'error': 'Use solo letras, numeros o guion en el modulo'}), 400
    if room_count <= 0:
        return jsonify({'error': 'La cantidad de habitaciones debe ser mayor a 0'}), 400
    if room_count > 500:
        return jsonify({'error': 'La cantidad de habitaciones no puede superar 500'}), 400
    if Room.query.filter_by(building=building).first():
        return jsonify({'error': 'Ese modulo ya existe'}), 400
    width = max(2, len(str(room_count)))
    new_rooms = []
    for number in range(1, room_count + 1):
        code = f'{building}{str(number).zfill(width)}'
        if Room.query.filter_by(code=code).first():
            return jsonify({'error': f'Ya existe una habitacion con codigo {code}'}), 400
        new_rooms.append(Room(code=code, building=building))
    db.session.add_all(new_rooms)
    db.session.commit()
    return jsonify({'ok': True, 'building': building, 'room_count': room_count}), 201


@app.route('/api/modules/<building>', methods=['DELETE'])
def api_delete_module(building):
    building = normalize_module_name(building)
    data = request.get_json(silent=True) or {}
    if data.get('password', '') != DELETE_PASSWORD:
        return jsonify({'error': 'Clave incorrecta'}), 403
    rooms = Room.query.filter_by(building=building).all()
    if not rooms:
        return jsonify({'error': 'El modulo no existe'}), 404
    room_ids = [r.id for r in rooms]
    deleted_cl = Checklist.query.filter(Checklist.room_id.in_(room_ids)).delete(synchronize_session=False)
    deleted_r = Room.query.filter(Room.id.in_(room_ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'ok': True, 'building': building, 'deleted_rooms': deleted_r, 'deleted_checklists': deleted_cl})


# ── API: Receptionists ────────────────────────────────────────────────────

@app.route('/api/receptionists', methods=['GET'])
def api_get_receptionists():
    recs = Receptionist.query.filter_by(active=True).order_by(Receptionist.name).all()
    return jsonify([r.to_dict() for r in recs])


@app.route('/api/receptionists', methods=['POST'])
def api_create_receptionist():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'El nombre es requerido'}), 400
    existing = Receptionist.query.filter_by(name=name).first()
    if existing:
        if not existing.active:
            existing.active = True
            db.session.commit()
            return jsonify(existing.to_dict()), 200
        return jsonify({'error': 'Ya existe un recepcionista con ese nombre'}), 400
    rec = Receptionist(name=name)
    db.session.add(rec)
    db.session.commit()
    return jsonify(rec.to_dict()), 201


@app.route('/api/receptionists/<int:rec_id>', methods=['PUT'])
def api_update_receptionist(rec_id):
    rec = Receptionist.query.get_or_404(rec_id)
    data = request.json
    if 'name' in data:
        rec.name = data['name'].strip()
    if 'active' in data:
        rec.active = data['active']
    db.session.commit()
    return jsonify(rec.to_dict())


@app.route('/api/receptionists/<int:rec_id>', methods=['DELETE'])
def api_delete_receptionist(rec_id):
    rec = Receptionist.query.get_or_404(rec_id)
    rec.active = False
    db.session.commit()
    return jsonify({'ok': True})


# ── API: Checklist ─────────────────────────────────────────────────────────

def checklist_from_data(data, room_id, receptionist_id):
    return Checklist(
        room_id=room_id,
        receptionist_id=receptionist_id,
        estado=data.get('estado', ''),
        luz_central=data.get('luz_central', ''),
        sensor=data.get('sensor', ''),
        cobertores=data.get('cobertores', ''),
        cambio_sabanas=data.get('cambio_sabanas', ''),
        velador=data.get('velador', ''),
        almohada=data.get('almohada', ''),
        extractor=data.get('extractor', ''),
        estufa=data.get('estufa', ''),
        basurero=data.get('basurero', ''),
        humidificador=data.get('humidificador', ''),
        cortina=data.get('cortina', ''),
        blackout=data.get('blackout', ''),
        aseo_general=data.get('aseo_general', ''),
        closet=data.get('closet', ''),
        observaciones=data.get('observaciones', '')
    )


@app.route('/api/checklist', methods=['POST'])
def api_create_checklist():
    data = request.json
    room_id = data.get('room_id')
    receptionist_id = data.get('receptionist_id')
    if not room_id or not receptionist_id:
        return jsonify({'error': 'room_id y receptionist_id son requeridos'}), 400
    checklist = checklist_from_data(data, room_id, receptionist_id)
    db.session.add(checklist)
    db.session.commit()
    return jsonify(checklist.to_dict()), 201


@app.route('/api/checklist/batch', methods=['POST'])
def api_create_checklist_batch():
    data = request.json
    items = data.get('items', [])
    receptionist_id = data.get('receptionist_id')
    if not items or not receptionist_id:
        return jsonify({'error': 'items y receptionist_id son requeridos'}), 400
    saved = []
    for item in items:
        cl = checklist_from_data(item, item['room_id'], receptionist_id)
        db.session.add(cl)
        saved.append(cl)
    db.session.commit()
    return jsonify({'saved': len(saved)}), 201


@app.route('/api/checklist/<int:checklist_id>', methods=['DELETE'])
def api_delete_checklist(checklist_id):
    data = request.json or {}
    if data.get('password', '') != DELETE_PASSWORD:
        return jsonify({'error': 'Clave incorrecta'}), 403
    checklist = Checklist.query.get_or_404(checklist_id)
    db.session.delete(checklist)
    db.session.commit()
    return jsonify({'ok': True})


# ── API: Dashboard ─────────────────────────────────────────────────────────

@app.route('/api/dashboard/stats')
def api_dashboard_stats():
    date_filter = get_dashboard_range()
    start_date = date_filter['start_date']
    end_date = date_filter['end_date']

    is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
    date_expr = func.date(Checklist.created_at) if is_sqlite else cast(Checklist.created_at, Date)

    base_filters = [Checklist.created_at >= start_date]
    if end_date is not None:
        base_filters.append(Checklist.created_at < end_date)

    daily = db.session.query(
        date_expr.label('date'), func.count(Checklist.id).label('count')
    ).filter(*base_filters).group_by(date_expr).order_by(date_expr).all()

    by_receptionist = db.session.query(
        Receptionist.name, func.count(Checklist.id).label('count')
    ).join(Checklist, Checklist.receptionist_id == Receptionist.id
    ).filter(*base_filters).group_by(Receptionist.name
    ).order_by(func.count(Checklist.id).desc()).all()

    issues_data = {}
    for field in OKX_FIELDS:
        count = Checklist.query.filter(*base_filters, getattr(Checklist, field) == 'x').count()
        issues_data[OKX_LABELS[field]] = count

    total_checklists = Checklist.query.filter(*base_filters).count()
    total_rooms = Room.query.count()
    rooms_checked = db.session.query(
        func.count(func.distinct(Checklist.room_id))
    ).filter(*base_filters).scalar()

    by_building = db.session.query(
        Room.building, func.count(Checklist.id).label('count')
    ).join(Checklist, Checklist.room_id == Room.id
    ).filter(*base_filters).group_by(Room.building).all()

    by_building_map = {row.building: row.count for row in by_building}
    ordered_buildings = sort_buildings(list(by_building_map.keys()))

    return jsonify({
        'daily': [{'date': str(d.date), 'count': d.count} for d in daily],
        'by_receptionist': [{'name': r.name, 'count': r.count} for r in by_receptionist],
        'issues': issues_data,
        'total_checklists': total_checklists,
        'total_rooms': total_rooms,
        'rooms_checked': rooms_checked or 0,
        'by_building': [{'building': b, 'count': by_building_map[b]} for b in ordered_buildings],
        'selected_date': date_filter['selected_date'],
        'days': date_filter['days'],
        'filter_label': date_filter['label']
    })


# ── API: History ───────────────────────────────────────────────────────────

@app.route('/api/history')
def api_history():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    building = request.args.get('building', '')
    receptionist_id = request.args.get('receptionist_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Checklist.query.join(Room).join(Receptionist)
    if building:
        query = query.filter(Room.building == building)
    if receptionist_id:
        query = query.filter(Checklist.receptionist_id == receptionist_id)
    if date_from:
        query = query.filter(Checklist.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Checklist.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))

    total = query.count()
    checklists = query.order_by(Checklist.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return jsonify({
        'items': [c.to_dict() for c in checklists],
        'total': total,
        'page': page,
        'pages': (total + per_page - 1) // per_page
    })


@app.route('/api/history/export')
def api_history_export():
    building = request.args.get('building', '')
    receptionist_id = request.args.get('receptionist_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Checklist.query.join(Room).join(Receptionist)
    if building:
        query = query.filter(Room.building == building)
    if receptionist_id:
        query = query.filter(Checklist.receptionist_id == receptionist_id)
    if date_from:
        query = query.filter(Checklist.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Checklist.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))

    checklists = query.order_by(Checklist.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Chequeos'

    headers = [
        'Fecha', 'Hora', 'Habitacion', 'Modulo', 'Recepcionista', 'Estado',
        'Luz Central', 'Sensor', 'Cobertores', 'Cambio Sabanas',
        'Velador', 'Almohada', 'Extractor', 'Estufa', 'Basurero',
        'Humidificador', 'Cortina', 'Blackout', 'Aseo General', 'Closet',
        'Observaciones'
    ]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ok_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    x_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    disp_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    ocup_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')

    for row_idx, c in enumerate(checklists, 2):
        values = [
            c.created_at.strftime('%Y-%m-%d'),
            c.created_at.strftime('%H:%M:%S'),
            c.room.code,
            c.room.building,
            c.receptionist.name,
            c.estado.capitalize() if c.estado else '',
            c.luz_central.upper() if c.luz_central else '',
            c.sensor.upper() if c.sensor else '',
            c.cobertores.upper() if c.cobertores else '',
            c.cambio_sabanas.upper() if c.cambio_sabanas else '',
            c.velador.upper() if c.velador else '',
            c.almohada.upper() if c.almohada else '',
            c.extractor.upper() if c.extractor else '',
            c.estufa.upper() if c.estufa else '',
            c.basurero.upper() if c.basurero else '',
            c.humidificador.upper() if c.humidificador else '',
            c.cortina.upper() if c.cortina else '',
            c.blackout.upper() if c.blackout else '',
            c.aseo_general.upper() if c.aseo_general else '',
            c.closet.upper() if c.closet else '',
            c.observaciones or ''
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            if 7 <= col_idx <= 20:
                if val == 'OK':
                    cell.fill = ok_fill
                elif val == 'X':
                    cell.fill = x_fill
            if col_idx == 6:
                lower = val.lower()
                if lower == 'disponible':
                    cell.fill = disp_fill
                elif lower == 'ocupada':
                    cell.fill = ocup_fill

    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(checklists) + 2, 100)):
            val = ws.cell(row=row, column=col).value
            if val and len(str(val)) > max_len:
                max_len = len(str(val))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 30)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:' + get_column_letter(len(headers)) + str(len(checklists) + 1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = 'chequeos_vca_' + now_santiago().strftime('%Y%m%d_%H%M') + '.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
